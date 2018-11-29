#coding=utf-8

from openpyxl import load_workbook
import shutil
import requests
import datetime
import os
from com.src.test.config import getConfig
from random import randint
import configparser
import re
import jsonpath
from com.src.test.logs import logger
import pymysql
from faker import Faker
import pandas
from Crypto.PublicKey import RSA
from Crypto.Cipher import PKCS1_v1_5 as Cipher_pkcs1_v1_5
import binascii
import base64
from com.src.test.encrpt import *
from requests.cookies import RequestsCookieJar


##初始化参数,把测试用例拷贝一份出来进行测试
def initParams():
    testcase_path = getConfig()["testcase_path"]
    testcase_name = getConfig()["testcase_name"]  ## 命名sheet名称以及测试名称
    now = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    results_path = os.path.join(os.path.abspath(os.path.pardir),'results')
    results_filename = os.path.join(results_path,'%s%s%s' % (now,testcase_name,os.path.splitext(testcase_path)[1]))
    shutil.copy(testcase_path,results_filename)
    return results_filename

def initLocustParams():
    testcase_path = getConfig()["testcase_path"]
    testcase_name = getConfig()["testcase_name"]  ## 命名sheet名称以及测试名称
    now = datetime.datetime.now().strftime('%Y%m%d')
    results_path = os.path.join(os.path.abspath(os.path.pardir),'results')
    results_filename = os.path.join(results_path,'%s%s%s' % (now,testcase_name,os.path.splitext(testcase_path)[1]))
    shutil.copy(testcase_path,results_filename)
    return results_filename

## 读取excel获取测试用例,返回数据格式为[[row1],[row2],[row3]]
def getTestData():
    case_result = initParams()
    wb = load_workbook(case_result)
    sheet = wb.active
    data_tuple = tuple(sheet.rows)[1:]
    parsedData = parseData(data_tuple)
    return parsedData

def getLocustTestData():
    case_result = initLocustParams()
    wb = load_workbook(case_result)
    sheet = wb.active
    data_tuple = tuple(sheet.rows)[1:]
    parsedData = parseLocustData(data_tuple)
    return parsedData

##将excel读取出的数据cell获取真正的数据返回格式为[[],[],[]]
def parseData(data_tuple):
    a_list = []
    for i in data_tuple:
        b_list = []
        for j in i:
            b_list.append(j.value)
        if b_list[9]:
            if b_list[9].strip().lower() == "yes":
                a_list.append(b_list)
    return a_list

def parseLocustData(data_tuple):
    a_list = []
    for i in data_tuple:
        b_list = []
        for j in i:
            b_list.append(j.value)
        if b_list[9]: ##那列
            if b_list[9].strip().lower() == "yes" and b_list[10].strip().lower() == "yes":
                a_list.append(b_list)
    return a_list

## 根据case数据进行接口请求
def requestAction(params):
    result = False
    params_dict = parseParams(params)
    # print("REQUEST BODY ：",params_dict)
    logger.info("#"*30)
    logger.info(params_dict)
    if params_dict["data"]:
        logger.info("POST_DATA")
        resp = requests.request(params_dict["method"],params_dict['url'],data=params_dict['data'],params=params_dict["payload"],cookies=params_dict["cookie"],headers= params_dict["headers"],verify=False)
    else:
        logger.info("GET_PARAMS")
        resp = requests.request(params_dict["method"],params_dict['url'],params=params_dict["payload"],cookies=params_dict["cookie"],headers= params_dict["headers"],verify=False)
    logger.info("respUrl:%s" % resp.url)
    logger.info("resp:%s" % resp.text)
    # print(resp.url)
    if resp.ok:
        resp_json = resp.json()

        # print("RESPJSON:",resp_json)
        store_resp = params_dict["store_resp"]

        ## 判断是否有需要保留的值，放入ini中
        ## 格式如下：{"tabooTest":["data","taboo"]}
        ## "tabooTest"为保存的变量名
        ## ["data","taboo"] ，为对应的key，data下面的taboo。。
        ## 修改为jsonpath格式的 {"jsonpath":"key"}
        ## jsonpath获取值，命名为key=值，放入ini
        if store_resp:
            ## jsonpath方式
            for k,v in store_resp.items():
                value = jsonpath.jsonpath(resp_json,k)
                # set_inis[v] = value
            # for k,v in set_inis.items():
                setRunningINI(v,value[0])
        ## 需要校验的预期结果，使用jsonpath，excel数据格式为{"jsonpath":value}
        check_point = params_dict['checks']
        if check_point:
            # print("check_point",check_point)
            logger.info("checkpoint:%s" % check_point)
            pass_num = 0
            for p in check_point:
                value = jsonpath.jsonpath(resp_json,p) ##resp_json 为需要处理的json，p为json路径的jsonpath
                # print("value",value[0])
                if value:
                    # logger.info(value)
                    # logger.info(value[0])
                    if value[0] == check_point[p]:
                        pass_num+=1
                else:
                    return result
            if len(check_point) == pass_num:
                logger.info("checkpoint校验通过！")
                result = True
        return result
    else:
        return result

##解析每一个case，params为一行数据
## 如果对应的value中有格式如下"${getRandomMobileNum}",需要调取getRandomMobileNum方法获取返回值，替换该值；
## 如果对应的value中有格式如下"$session",需要从running.ini中读取，session的值替换该值。
def parseParams(params):
    ## url
    url = ""
    path_url = ""
    host_url =""
    method =""
    headers={}
    cookie={}
    data={}
    payload={}
    checks={}
    desc="默认描述desc"
    exec="No"
    locust_exec="No"
    store_resp={}
    logger.info(params)
    if params[0] and params[1]: ## url
        url = params[0].strip() + params[1].strip()
    if params[0]:
        host_url = params[0].strip()
    if params[1]:
        path_url = params[1].strip()
    ## method
    if params[2]:
        method = params[2].strip()
    ##headers
    if params[3]:
        headers = eval(params[3].strip())
        headers = handleUSD(headers)

    ## cookie
    if params[4]:
        cookie = eval(params[4].strip())
        cookie = handleUSD(cookie)

    ## data
    elif params[5]:
        data = eval(params[5].strip())
        data = handleUSD(data)
    ##params payload
    if params[6]:
        payload = eval(params[6].strip())
        payload = handleUSD(payload)
    ##checks
    if params[7]:
        checks = eval(params[7].strip())

    ##description
    if params[8]:
        desc = params[8].strip()

    ##exec
    if params[9]:
        exec = params[9].strip()

    ##locust_exec
    if params[10]:
        locust_exec = params[10].strip()

    ##store_resp,格式如下{"tabooTest":["key1","key2,"key3"]
    # tabooTest为需要保存的key，保存的值为key1,key2,key3,的值，resp[key1][key2][key3]
    # 保存ini为tabooTest=获取到的值
    if params[11]:
        store_resp = eval(params[11].strip())

    return {"url":url,"host_url":host_url,"path_url":path_url,"method":method,"headers":headers,"cookie":cookie,"data":data,"payload":payload,"checks":checks,"desc":desc,"exec":exec,"locust_exec":locust_exec,"store_resp":store_resp}

def handleUSD(dict_obj):
    if isinstance(dict_obj,dict):
        for k,v in dict_obj.items():
            # print(k,v)
            if isinstance(v,dict):
                # for i,j in v:
                #     if isinstance(j,dict):
                #         handleUSD(v,dict_obj_tmp=dict_obj)
                handleUSD(v)
            elif isinstance(v,str) and v.startswith("${") and v.endswith("}"):
                matchObj = re.match(r"^\$\{(.*?)\((.*?)\)\}$",v)
                meth = matchObj.group(1)
                trans_prams = matchObj.group(2)
                if meth and trans_prams:
                    dict_obj[k]=eval(meth)(trans_prams)
                    # print(eval(meth)(trans_prams))
                elif meth:
                    dict_obj[k] = eval(meth)()
                    # print(eval(meth))
                # print("@@",meth,trans_prams)
            elif isinstance(v,str) and v.startswith("$"):
                ini_param = v[1:]
                dict_obj[k] = getRunningINI(ini_param)
                # print("##",getRunningINI("ppp"))
                # print("!!",ini_param)
    ## 处理整体生成请求参数，如果格式json，data，params字段符合格式"${method()}",请求结果替换为整体入参。
    elif isinstance(dict_obj,str) and dict_obj.startswith("${") and dict_obj.endswith("}"):
        matchObj = re.match(r"^\$\{(.*?)\((.*?)\)\}$",dict_obj)
        meth = matchObj.group(1)
        trans_prams = matchObj.group(2)
        if meth and trans_prams:
            dict_obj=eval(meth)(trans_prams)
            # print(eval(meth)(trans_prams))
        elif meth:
            dict_obj = eval(meth)()
    return dict_obj

def setRunningINI(s_key,s_value,section="running",encoding="utf-8"):
    logger.info("准备写入%s{%s:%s}" % (section, s_key, s_value))
    if not isinstance(s_value,str):
        s_value = str(s_value)
    if not isinstance(s_key,str):
        s_key = str(s_key)
    cf = configparser.ConfigParser()
    res = cf.read("running.ini",encoding=encoding)
    logger.error(res)
    if not cf.has_section(section):
        cf.add_section(section)
        logger.info("添加section :%s" % section)
    cf.set(section,s_key,s_value)
    with open("running.ini","w") as f2:
        cf.write(f2)
    logger.info("写入成功%s{%s:%s}" % (section, s_key, s_value))

def getRunningINI(s_key,section="running",inifile="running.ini"):
    cf = configparser.ConfigParser()
    cf.read(inifile)
    res = cf.get(section,s_key)
    if not section == "email":
        logger.info("从%s中读取%s,值为%s" % (section,s_key,res))
    return res

def getRandomMobileNum(returnStr=True):
    mobileNo = randint(13000000000,13999999999)
    logger.info("获取到随机手机号码：%s" % mobileNo)
    if returnStr:
        return str(mobileNo)
    return mobileNo

def getRandomPort():
    mobileNo = randint(9000,50000)
    return mobileNo

def getOneIP():
    return "99.99.99.98"


def getCurrentTimeHour():
    return datetime.datetime.now().strftime("%Y%m%d%H")

def getCurrentTimeDay():
    return datetime.datetime.now().strftime("%Y%m%d")

def getConnCursor():
    conn = pymysql.connect(getRunningINI("host", "mysql"), getRunningINI("user", "mysql"),
                           getRunningINI("passwd", "mysql"), getRunningINI("database", "mysql"))
    cursor = conn.cursor()
    return conn,cursor

def getMobile():
    conn,cursor = getConnCursor()
    cursor.execute("select mobile from users")
    res = cursor.fetchOne()
    cursor.close()
    conn.close()
    return res

def writeTestData(s_key, s_value, section="data"):
    conn,cursor = getConnCursor()
    cursor.execute()
    res = cursor.fetchall()
    cursor.close()
    conn.close()
    cf = configparser.ConfigParser()
    if not cf.has_section(section):
        cf.add_section(section)
    for i in range(res):
        cf.set(section, i, res[i])
        logger.info("写入%s{%s:%s}" % (section, s_key, s_value))
        with open("testdata.ini", "w") as f1:
            cf.write(f1)

def getFakePhoneNum():
    faker = Faker("zh_CN")
    return faker.phone_number()

def getIDCARD():
    faker = Faker("zh_CN")
    name = faker.name()
    logger.info(name)
    return name

def genPhoneNumFromCSV():
    csv_data = pandas.read_csv("testdata.csv",header=None)
    for i in range(len(csv_data)):
        yield csv_data.iloc[i,0]
def genPasswdFromCSV():
    csv_data = pandas.read_csv("testdata.csv",header=None)
    for i in range(len(csv_data)):
        yield csv_data.iloc[i,1]

def genPhoneAndPasswdFromCSV():
    csv_data = pandas.read_csv("testdata.csv",header=None)
    for i in range(len(csv_data)):
        yield csv_data.iloc[i,0],csv_data.iloc[i,1]

phoneNum = genPhoneNumFromCSV()
passwd = genPasswdFromCSV()
phoneAndPasswd = genPhoneAndPasswdFromCSV()

def nextPhone():
    try:
        global phoneNum
        phone = next(phoneNum)
        logger.info("phone:%s" % phone)
        return phone
    except StopIteration as e:
        logger.error("StopIteration：%s" % e)
        phoneNum = genPhoneNumFromCSV()
        logger.info("重新创建生成器：%s" % "genPhoneNumFromCSV")
        phone = next(phoneNum)
        logger.info("phone:%s" % phone)
        return phone
def nextPasswd():
    try:
        global passwd
        passwd_next = next(passwd)
        logger.info("passwd_next:%s" % passwd_next)
        return passwd
    except StopIteration as e:
        logger.error("StopIteration：%s" % e)
        passwd = genPhoneNumFromCSV()
        logger.info("重新创建生成器：%s" % "genPasswdFromCSV")
        passwd_next = next(passwd)
        logger.info("passwd_next:%s" % passwd_next)
        return passwd

def nextPhoneAndPasswd():
    try:
        global phoneAndPasswd
        phoneAndPasswd_next = next(phoneAndPasswd)
        logger.info("phoneAndPasswd_next:%s,%s" % phoneAndPasswd_next)
        return phoneAndPasswd_next
    except StopIteration as e:
        logger.error("StopIteration：%s" % e)
        phoneAndPasswd = genPhoneAndPasswdFromCSV()
        logger.info("重新创建生成器：%s" % "genPhoneAndPasswdFromCSV")
        phoneAndPasswd_next = next(phoneAndPasswd)
        return phoneAndPasswd_next

def getIP():
    faker = Faker("zh_CN")
    ip = faker.ipv4()
    logger.info("随机生成ip：%s" % ip)
    return ip


def setCookieNEO_SES(key):
    cookie_jar = RequestsCookieJar()
    cookie_jar.set(key, getRunningINI(key), path="/",domain=".dongmanmanhua.cn")
    return cookie_jar



def getencpw():
    phone= "15010173763"
    passwd = "123qwe"
    resp = requests.post(" http://dev.dongmanmanhua.cn/member/login/rsa/getKeys")
    resp_json = resp.json()
    sessionKey = resp_json["sessionKey"]
    keyName = resp_json["keyName"]
    nvalue = int(resp_json["nvalue"])
    evalue = int(resp_json["evalue"],16)
    key =  RSA.construct((evalue,nvalue)) #根据e,n生成publicKey
    logger.info(key)

    text = chr(len(sessionKey))+sessionKey+chr(len(phone))+phone+chr(len(passwd))+passwd
    # encpw = rsa.encrypt(text,key)
    # text="www.baidu.com"
    text = text.encode("utf-8")
    cipher = Cipher_pkcs1_v1_5.new(key)
    encpw = cipher.encrypt(text)
    base64.b64encode(encpw)
    encpw = binascii.b2a_hex(encpw)
    return keyName,encpw



def getLoginForms():
    phone,passwd = nextPhoneAndPasswd()
    # logger.info("请输入验证码：")
    # passwd = input("请输入验证码：")
    return getLoin(str(phone),str(passwd))





if __name__ == "__main__":
    # print(getTestData())
    # dict_obj = {"q1":"123","q2":"${getRandomMobileNum()}","q3":"${getRandomMobileNum(33)}","q4":"$ppp"}
    # dict_obj2 = {"q1": "123", "q2": "${getRandomMobileNum()}", "q3": "${getRandomMobileNum(33)}", "q4": "$ppp","q5":{"q6":"${getRandomMobileNum(33)}"}}
    # print(handleUSD(dict_obj2))
    # print(getIDCARD())
    print(getencpw())
