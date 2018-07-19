from openpyxl import load_workbook,Workbook
import requests
import json
import datetime
import shutil
import os
import xlrd
from xlutils.copy import copy
from pymongo import MongoClient
import base64
# import pytesseract
# from PIL import Image



### 接口地址	endpoint	method	headers	json	data	params	checkpoint	scene_desc	exec	resp_body	result   database            dynamic_params
###    0       1            2     3       4       5       6         7          8          9         10        11       12          13   14

### results 格式为 [[row,resp-body,pass-or-fail],]

testcase_path = 'testcase.xlsx'
# testcase_path = 'testcase.xls'
testcase_name = '天行数科stb'  ## 命名sheet名称以及测试名称
now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
results_path = os.path.join(os.path.abspath(os.path.pardir),'results')
results_filename = os.path.join(results_path,'%s%s%s' % (now,testcase_name,os.path.splitext(testcase_path)[1]))
share_params = {}


def readCase(case):
    shutil.copy(case,results_filename)
    results =[]
    wb = load_workbook(results_filename)
    sheet = wb.active
    sheet.title = testcase_name
    data_tuple = tuple(sheet.rows)[1:]
    pass_count = 0
    action_count = 0
    for row in data_tuple:  ##过滤掉标题，从第二行开始取数据
        row_data = tuple(map(lambda x:x.value,row))
        result = startRequest(data_tuple.index(row)+2,row_data)
        if result:
            action_count += 1
            if result[2]:
                pass_count += 1
            results.append(result)
    writeResults(wb,results)
    print('执行用例数：%s，通过%s' % (action_count,pass_count))

def startRequest(row,data):  ## 返回 [row,resp-body,pass-or-fail]
    toChangeParams = {}
    req_yes = data[9]
    if req_yes and req_yes.lower() == 'yes':
        req_method = data[2]
        req_url = data[0] + data[1]
        req_check  = eval(data[7])
        print(data[8])
        ###  如果有动态参数就处理
        dynamic_params = data[14]
        if dynamic_params:
            dp = eval(dynamic_params)
            #######   替换有current的 参数
            if 'current' in dp:
                current_data = dp['current']
                if isinstance(current_data,list):
                    for i in current_data:
                        toChangeParams[i] = share_params[i]
                else:
                    pass ##  有问题什么都不干

        if req_method == "POST":
            if data[4]:
                if toChangeParams:
                    request_data = eval(data[4])
                    for key in toChangeParams:
                        if "params" in request_data:
                            request_data['params'][key] =  toChangeParams[key]
                        else:
                            request_data[key] = toChangeParams[key]
                    print(row,"@@",request_data)
                    resp = requests.request(req_method, req_url,headers=eval(data[3]),json=request_data)
                else:
                    print(row,"@@",eval(data[4]))
                    resp = requests.request(req_method, req_url,headers=eval(data[3]),json=eval(data[4]))

            elif data[5]:
                if toChangeParams:
                    request_data = eval(data[5])
                    for key in toChangeParams:
                        request_data['params'][key] =  toChangeParams[key]
                    print(row,"@@",request_data)
                    resp = requests.request(req_method, req_url,headers=eval(data[3]), data=request_data)
                else:
                    # print(row, "@@", eval(data[5]))
                    resp = requests.request(req_method, req_url, headers=eval(data[3]), data=eval(data[5]))

            else:
                resp = requests.request(req_method, req_url,headers=eval(data[3]), data={})

        elif req_method == "GET":
            if toChangeParams:
                request_data = eval(data[6])
                for key in toChangeParams:
                    request_data['params'][key] = toChangeParams[key]
                resp = requests.request(req_method, req_url,headers=eval(data[3]), params=request_data)
            else:
                resp = requests.request(req_method, req_url,headers=eval(data[3]), params=eval(data[6]))

        else:
            resp = "" ## 目前仅支持GET和POST请求！
        print(row,"##",resp.text)


        return handleResp(row,resp,req_check,data)
    return False


def handleResp(row,resp,cp,data): ## 返回 [row,resp-body,pass-or-fail]
    result = False
    if resp:
        try:
            resp_json = resp.json()
            result = checkPoint(cp, resp_json,data)

            ### 将需要保存的参数提取
            dynamic_params = data[14]
            if dynamic_params:
                dp = eval(dynamic_params)
                if "next" in dp:
                    next_data = dp['next']
                    if isinstance(next_data, list):
                        for i in next_data:
                            resp_i = resp_json[i[0]]
                            # print(resp_i)
                            if i[2] == "SAME":
                                share_params[i[1]] = resp_i
                            elif i[2] == 'IMAGEBASE64':
                                base64ToImage(resp_i)
                                autoCode = input("请输入图形验证码：")
                                share_params[i[1]] = autoCode
                    elif isinstance(next_data, dict):
                        for key in next_data:
                            value = next_data[key]
                            if isinstance(value, list):
                                for i in value:
                                    resp_i = resp_json[key][i[0]]
                                    # print(resp_i)
                                    if i[2] == "SAME":
                                        share_params[i[1]] = resp_i
                                    elif i[2] == 'IMAGEBASE64':
                                        png = base64ToImage(resp_i)
                                        # code = ocrImage(png)
                                        autoCode = input("请输入图形验证码：")
                                        share_params[i[1]] = autoCode

                            elif isinstance(value, dict):
                                for inkey in value:
                                    invalue = value[inkey]
                                    if isinstance(invalue, list):
                                        for ini in invalue:
                                            resp_i = resp_json[key][inkey][ini[0]]
                                            if ini[2] == "SAME":
                                                share_params[ini[1]] = resp_i
                                            elif ini[2] == 'IMAGEBASE64':
                                                base64ToImage(resp_i)
                                                autoCode = input("请输入图形验证码：")
                                                share_params[ini[1]] = autoCode

                    else:
                        pass  ##  有问题什么都不干
            json_string = json.dumps(resp_json,indent=4,ensure_ascii=False)
            # if len(json_string)> 32737:
            #     with open("%s%s.log" % (now+testcase_name,row),"w") as f1:
            #         f1.write(json_string)
            #     resp= '响应值过大，请参看日志记录'
            # else:
            #     resp = json_string
            resp = json_string
        except ValueError:
            resp = resp.text
    else:
        resp = "目前仅支持GET和POST请求！"
    return row,resp,result

## 传入 检查点，响应json，如果有不是dict的，直接校验失败
def checkPoint(checkpoint,resp_json,data):  ## 返回成功 True，失败False
    if checkpoint:
        if isinstance(resp_json, dict) and isinstance(checkpoint, dict):
            success_count = 0
            for p in checkpoint:
                check_data = checkpoint[p]
                if isinstance(check_data, dict):
                    inner_success_count = 0
                    for inner in check_data:
                        p_resp_json = resp_json.get(p,None)
                        if p_resp_json and isinstance(p_resp_json,dict):
                            if check_data[inner] == p_resp_json.get(inner,None):
                                inner_success_count += 1
                            elif check_data[inner].startswith('LIKE'):
                                if check_data[inner][4:] in p_resp_json.get(inner, ''):
                                    inner_success_count += 1
                    if len(check_data) == inner_success_count:
                        success_count += 1

                elif isinstance(check_data,str):

                    if check_data == resp_json.get(p,None):
                        success_count += 1
                    elif check_data.startswith('LIKE'):
                        if check_data[4:] in resp_json.get(p,''):
                            success_count += 1

                elif check_data is None:
                    if check_data == resp_json.get(p,True):
                        success_count += 1
                else:
                    if check_data == resp_json.get(p,True):
                        success_count += 1
            if len(checkpoint) == success_count:
                if data[12]:
                    db_check = eval(data[12])
                    if "mongo" in db_check:
                        mongo_data = db_check['mongo']
                        db_result = checkMongo(mongo_data,resp_json[mongo_data['key']])
                        return db_result

                return True
            else:
                return False
        else:
            return False
    else :
        return True

def writeResults(wb,results): ## 11写入body，12写入pass or fail  [row,body,pass]
    sheet = wb.active
    for i in results:

        sheet.cell(i[0],11,i[1])
        if isinstance(i[2],dict):
            sheet.cell(i[0],14,str(i[2]))
            sheet.cell(i[0], 12, True)
        else:
            sheet.cell(i[0], 12, i[2])
    wb.save(results_filename)
    wb.close()


######-----------以上处理xlsx--------##############
######------------------------------##############
######------------------------------##############
######------------------------------##############
######------------------------------##############
######-----------以下处理xls---------##############



pass_count = 0
case_count = 0


def doTest(testcase):
    print("####----    执行中...")
    shutil.copy(testcase,results_filename)
    data = xlrd.open_workbook(results_filename)
    table = data.sheet_by_index(0)
    write_data = copy(data)
    for row in range(1, table.nrows):
        callAPI(os.path.splitext(testcase)[0],write_data, row, tuple(table.row_values(row)))
    write_data.save(results_filename)
    print("####----    总共%s个CASE,通过%s个 CASE" % (case_count, pass_count))

def callAPI(testcase_name,write_data, row, params):
    if params[9] == 'yes':
        global case_count
        case_count += 1
        resp = ''
        resp_json = ''
        if params[2] == "POST":
            if params[4]:
                resp = requests.request(params[2], params[0] + params[1], headers=eval(params[3]), json=eval(params[4]))

            elif params[5]:
                resp = requests.request(params[2], params[0] + params[1], headers=eval(params[3]), data=eval(params[5]))
        elif params[2] == "GET":
            resp = requests.request(params[2], params[0] + params[1], headers=eval(params[3]), params=eval(params[6]))

        print(row + 1, "RESP::", resp.text)
        try:
            resp_json = resp.json()
            json_string = json.dumps(resp_json, indent=4, ensure_ascii=False)
            if len(json_string) > 32737:
                with open("%s%s.txt" % (testcase_name, row + 1), "w") as f1:
                    f1.write(json_string.encode("utf-8"))
                json_string = u'响应值过大，请参看日志记录'
            write_data.get_sheet(0).write(row, 10, json_string)
        except ValueError:
            resp_text = resp.text
            write_data.get_sheet(0).write(row, 10, resp_text)

        pass_res = "FAIL"

        if isinstance(resp_json, dict) and params[7]:
            pass_num = 0
            check_point_dict = eval(params[7])
            for i in check_point_dict:
                check_data = check_point_dict[i]
                if isinstance(check_data, dict):
                    inner_pass_num = 0
                    for j in check_data:
                        try:
                            data_value = resp_json[i][j]
                            if isinstance(check_data[j], str) and check_data[j].startswith('LIKE'):
                                assert check_data[j][4:] in data_value
                                inner_pass_num += 1
                            else:
                                assert check_data[j] == resp_json[i][j]
                                inner_pass_num += 1
                        except Exception:
                            break
                    if len(check_data) == inner_pass_num:
                        pass_num += 1

                else:
                    if isinstance(check_data, str) and check_data.startswith('LIKE'):
                        if check_data[4:] in resp_json.get(i, None):
                            pass_num += 1
                    else:
                        if resp_json.get(i, None) == check_data:
                            pass_num += 1

            check_point_dict_length = len(check_point_dict)
            if pass_num == check_point_dict_length:
                global pass_count
                # print('@@@@@@@@@@@@',params[12])
                if params[12]:
                    # print('#################################')
                    db_check = eval(params[12])
                    if "mongo" in db_check:
                        mongo_data = db_check['mongo']
                        db_result = checkMongo(mongo_data,resp_json[mongo_data['key']])
                        if db_result:
                            write_data.get_sheet(0).write(row, 13, str(db_result))
                            pass_res = 'PASS'
                            pass_count += 1
                else:
                    pass_res = 'PASS'
                    pass_count += 1
        write_data.get_sheet(0).write(row, 11, pass_res)



##### ------############### 公用方法 #############################################

def checkMongo(data,uuid):
    # print("#"*20,uuid)
    conn = MongoClient(data['host'],data['port'])
    db = conn[data['database']]
    if db.authenticate(data['username'],data['password']):
        table = db[data['table']]
        result = table.find_one({data['key']:uuid})
        print("@"*20,uuid,result)
        if result:
            conn.close()
            return result
        else:
            conn.close()
            return False
    else:
        conn.close()
        return False



def base64ToImage(base64_text):
    image_data = base64.b64decode(base64_text)
    tmp1_png = os.path.join(os.path.dirname(__file__), "..","tmp","tmp.png")
    tmp1 = open(tmp1_png, "wb")
    tmp1.write(image_data)
    tmp1.close()
    return os.path.abspath(tmp1_png)


def ocrImage(png):
    image = Image.open(png)
    code =  pytesseract.image_to_string(image)
    return code



if __name__ == '__main__':
    extension_name = os.path.splitext(testcase_path)[1]
    if extension_name == '.xlsx':
        readCase(testcase_path)
    elif extension_name == '.xls':
        doTest(testcase_path)
    else:
        print('该脚本不支持的文件格式，请检查！')